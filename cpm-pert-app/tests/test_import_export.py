"""
E2E tests for import (CSV / JSON / XLSX) and export (Gantt PNG / Network PNG).

Import tests use a fresh isolated page per test so state never bleeds between them.
Export and UX tests share an already-analyzed page (module-scoped fixture).

Requires a running Flask server at http://127.0.0.1:5000.
"""

import json
import re

import openpyxl
import pytest
from playwright.sync_api import expect

from conftest import BASE_URL, fill_rows, click_analyze_and_capture

# ── Shared analyzed page (module-scoped) ─────────────────────────────────────

_shared_ctx = None
_shared_page = None

ANALYZED_ROWS = [
    {"id": "A", "name": "Task A", "duration": "5", "dependencies": ""},
    {"id": "B", "name": "Task B", "duration": "3", "dependencies": "A"},
    {"id": "C", "name": "Task C", "duration": "4", "dependencies": "A"},
]


@pytest.fixture(scope="module")
def analyzed_page(browser):
    """Fresh page with a 3-task CPM analysis already run — shared across UX / export tests."""
    global _shared_ctx, _shared_page
    _shared_ctx = browser.new_context()
    _shared_page = _shared_ctx.new_page()
    _shared_page.goto(BASE_URL, wait_until="domcontentloaded")
    fill_rows(_shared_page, ANALYZED_ROWS)
    resp, _ = click_analyze_and_capture(_shared_page)
    assert resp.status == 200
    _shared_page.wait_for_selector(".gantt-row")
    yield _shared_page
    _shared_ctx.close()


# ── Group 1: Import UX ────────────────────────────────────────────────────────

def test_import_dropdown_shows_three_format_options(analyzed_page):
    """Opening the Import dropdown reveals CSV, JSON and Excel options."""
    page = analyzed_page
    page.locator("button.dropdown-toggle", has_text="Import").click()

    menu = page.locator(".dropdown-menu")
    expect(menu).to_be_visible()
    expect(menu.locator("label[for='file-upload-csv']")).to_be_visible()
    expect(menu.locator("label[for='file-upload-json']")).to_be_visible()
    expect(menu.locator("label[for='file-upload-xlsx']")).to_be_visible()

    page.keyboard.press("Escape")


def test_format_modal_opens_and_shows_all_formats(analyzed_page):
    """Clicking the ⓘ button opens the format reference modal with CSV/JSON/Excel docs."""
    page = analyzed_page
    page.locator("button[data-bs-target='#import-format-modal']").click()

    modal = page.locator("#import-format-modal")
    expect(modal).to_be_visible()
    expect(modal.locator(".modal-title")).to_contain_text("Import File Formats")

    body = modal.locator(".modal-body")
    expect(body).to_contain_text("CSV")
    expect(body).to_contain_text("JSON")
    expect(body).to_contain_text("Excel")
    expect(body).to_contain_text("ac,pr,du,name")  # canonical CSV header example

    # Close via the footer Close button (X icon also has data-bs-dismiss — need specific selector)
    modal.locator(".modal-footer button").click()
    expect(modal).not_to_be_visible()


# ── Group 2: CSV import ───────────────────────────────────────────────────────

def test_csv_import_populates_table(page, tmp_path):
    """Uploading a valid CSV file populates the input table with the correct rows."""
    csv_file = tmp_path / "tasks.csv"
    csv_file.write_text("ac,pr,du,name\nA,-,3,Design\nB,A,5,Build\nC,AB,2,Test\n", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-csv").set_input_files(str(csv_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(3)

    # Row 1 — A, duration 3
    expect(rows.nth(0).locator("td:nth-child(1)")).to_have_text("A")
    expect(rows.nth(0).locator("td:nth-child(3)")).to_have_text("3")

    # Row 3 — C, depends on A and B (parseCsvPredecessors("AB") → ["A","B"])
    expect(rows.nth(2).locator("td:nth-child(1)")).to_have_text("C")
    dep_text = rows.nth(2).locator("td:nth-child(4)").text_content()
    assert "A" in dep_text and "B" in dep_text


def test_csv_pert_import_populates_pert_columns(page, tmp_path):
    """Uploading a PERT CSV with opt/ml/pess headers populates O/M/P columns and auto-switches to PERT mode."""
    csv_file = tmp_path / "pert.csv"
    csv_file.write_text(
        "ac,pr,opt,ml,pess,name\nA,-,2,4,6,Alpha\nB,A,1,3,5,Beta\n",
        encoding="utf-8",
    )
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-csv").set_input_files(str(csv_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(2)

    expect(page.locator("#toggle-pert")).to_be_checked()

    row1 = rows.nth(0)
    expect(row1.locator("td:nth-child(5)")).to_be_visible()
    expect(row1.locator("td:nth-child(5)")).to_have_text("2")  # O
    expect(row1.locator("td:nth-child(6)")).to_have_text("4")  # M
    expect(row1.locator("td:nth-child(7)")).to_have_text("6")  # P

    row2 = rows.nth(1)
    expect(row2.locator("td:nth-child(5)")).to_have_text("1")
    expect(row2.locator("td:nth-child(6)")).to_have_text("3")
    expect(row2.locator("td:nth-child(7)")).to_have_text("5")


def test_csv_pert_import_auto_switches_to_pert_mode(page, tmp_path):
    """Uploading a PERT CSV while in CPM mode automatically flips the mode toggle to PERT."""
    csv_file = tmp_path / "pert.csv"
    csv_file.write_text("ac,pr,opt,ml,pess\nA,-,2,4,6\n", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    expect(page.locator("#toggle-pert")).not_to_be_checked()

    page.locator("#file-upload-csv").set_input_files(str(csv_file))

    expect(page.locator("#toggle-pert")).to_be_checked()
    expect(page.locator("#input-table tbody tr")).to_have_count(1)


def test_csv_cpm_import_while_in_pert_mode_auto_switches(page, tmp_path):
    """Uploading a CPM CSV while in PERT mode automatically flips the mode toggle back to CPM."""
    csv_file = tmp_path / "cpm.csv"
    csv_file.write_text("ac,pr,du\nA,-,3\nB,A,5\n", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#toggle-pert").check()
    expect(page.locator("#toggle-pert")).to_be_checked()

    page.locator("#file-upload-csv").set_input_files(str(csv_file))

    expect(page.locator("#toggle-pert")).not_to_be_checked()
    expect(page.locator("#input-table tbody tr")).to_have_count(2)


def test_csv_pert_wrong_aliases_shows_error(page, tmp_path):
    """CSV with long-form PERT aliases (optimistic/most_likely/pessimistic) is rejected with an explicit column error."""
    csv_file = tmp_path / "wrong_pert.csv"
    csv_file.write_text(
        "ac,pr,optimistic,most_likely,pessimistic,name\nA,-,2,4,6,Alpha\n",
        encoding="utf-8",
    )
    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-csv").set_input_files(str(csv_file))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("Unrecognised column")


def test_csv_import_allows_reanalysis(page, tmp_path):
    """After a CSV import the Analyze button produces a valid CPM result."""
    csv_file = tmp_path / "tasks.csv"
    csv_file.write_text("ac,pr,du\nA,-,4\nB,A,6\n", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-csv").set_input_files(str(csv_file))
    page.locator("#input-table tbody tr").nth(0)  # wait for table

    resp, _ = click_analyze_and_capture(page)
    assert resp.status == 200
    page.wait_for_selector(".gantt-row")
    assert page.locator(".gantt-row").count() == 2


# ── Group 3: JSON import ──────────────────────────────────────────────────────

def test_json_cpm_import_populates_table(page, tmp_path):
    """Uploading a JSON file with CPM task objects populates the table correctly."""
    tasks = [
        {"id": "A", "name": "Alpha", "duration": 4, "dependencies": ""},
        {"id": "B", "name": "Beta",  "duration": 7, "dependencies": "A"},
    ]
    json_file = tmp_path / "tasks.json"
    json_file.write_text(json.dumps(tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(json_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(2)
    expect(rows.nth(0).locator("td:nth-child(1)")).to_have_text("A")
    expect(rows.nth(0).locator("td:nth-child(3)")).to_have_text("4")
    expect(rows.nth(1).locator("td:nth-child(1)")).to_have_text("B")
    expect(rows.nth(1).locator("td:nth-child(3)")).to_have_text("7")
    expect(rows.nth(1).locator("td:nth-child(4)")).to_have_text("A")


def test_json_pert_import_populates_pert_columns(page, tmp_path):
    """Uploading a JSON file with PERT fields populates O/M/P columns and auto-switches to PERT mode."""
    tasks = [
        {"id": "A", "name": "Alpha", "optimistic": 2, "most_likely": 4, "pessimistic": 6, "dependencies": ""},
        {"id": "B", "name": "Beta",  "optimistic": 1, "most_likely": 3, "pessimistic": 5, "dependencies": "A"},
    ]
    json_file = tmp_path / "pert_tasks.json"
    json_file.write_text(json.dumps(tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(json_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(2)

    expect(page.locator("#toggle-pert")).to_be_checked()

    # Columns (1-based nth-child): ID=1, Name=2, Dur=3, Dep=4, O=5, M=6, P=7
    row1 = rows.nth(0)
    expect(row1.locator("td:nth-child(5)")).to_be_visible()
    expect(row1.locator("td:nth-child(5)")).to_have_text("2")
    expect(row1.locator("td:nth-child(6)")).to_have_text("4")
    expect(row1.locator("td:nth-child(7)")).to_have_text("6")

    row2 = rows.nth(1)
    expect(row2.locator("td:nth-child(5)")).to_have_text("1")
    expect(row2.locator("td:nth-child(6)")).to_have_text("3")
    expect(row2.locator("td:nth-child(7)")).to_have_text("5")


def test_json_pert_import_auto_switches_to_pert_mode(page, tmp_path):
    """Uploading a PERT JSON file while in CPM mode automatically flips the mode toggle to PERT."""
    tasks = [{"id": "A", "optimistic": 2, "most_likely": 4, "pessimistic": 6, "dependencies": ""}]
    json_file = tmp_path / "pert.json"
    json_file.write_text(json.dumps(tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    expect(page.locator("#toggle-pert")).not_to_be_checked()

    page.locator("#file-upload-json").set_input_files(str(json_file))

    expect(page.locator("#toggle-pert")).to_be_checked()
    expect(page.locator("#input-table tbody tr")).to_have_count(1)


def test_json_cpm_import_while_in_pert_mode_auto_switches(page, tmp_path):
    """Uploading a CPM JSON file while in PERT mode automatically flips the toggle back to CPM."""
    tasks = [
        {"id": "A", "name": "Alpha", "duration": 3, "dependencies": ""},
        {"id": "B", "name": "Beta",  "duration": 5, "dependencies": "A"},
    ]
    json_file = tmp_path / "cpm.json"
    json_file.write_text(json.dumps(tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#toggle-pert").check()
    expect(page.locator("#toggle-pert")).to_be_checked()

    page.locator("#file-upload-json").set_input_files(str(json_file))

    expect(page.locator("#toggle-pert")).not_to_be_checked()
    expect(page.locator("#input-table tbody tr")).to_have_count(2)


def test_json_import_with_array_dependencies(page, tmp_path):
    """JSON tasks may use an array for dependencies — they are joined to a comma string."""
    tasks = [
        {"id": "A", "name": "Alpha", "duration": 3, "dependencies": []},
        {"id": "B", "name": "Beta",  "duration": 2, "dependencies": ["A"]},
        {"id": "C", "name": "Gamma", "duration": 4, "dependencies": ["A", "B"]},
    ]
    json_file = tmp_path / "deps.json"
    json_file.write_text(json.dumps(tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(json_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(3)
    dep_c = rows.nth(2).locator("td:nth-child(4)").text_content()
    assert "A" in dep_c and "B" in dep_c


# ── Group 4: XLSX import ──────────────────────────────────────────────────────

def test_xlsx_import_populates_table(page, tmp_path):
    """Uploading a valid Excel file populates the input table correctly."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ac", "pr", "du", "name"])
    ws.append(["A", "-", 3, "Design"])
    ws.append(["B", "A", 5, "Build"])
    ws.append(["C", "AB", 2, "Test"])
    xlsx_file = tmp_path / "tasks.xlsx"
    wb.save(str(xlsx_file))

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-xlsx").set_input_files(str(xlsx_file))

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(3)
    expect(rows.nth(0).locator("td:nth-child(1)")).to_have_text("A")
    expect(rows.nth(0).locator("td:nth-child(3)")).to_have_text("3")
    expect(rows.nth(2).locator("td:nth-child(1)")).to_have_text("C")


def test_xlsx_pert_import_populates_table(page, tmp_path):
    """Uploading a PERT Excel file auto-switches to PERT mode and populates O/M/P columns."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["ac", "pr", "opt", "ml", "pess", "name"])
    ws.append(["A", "-", 2, 4, 6, "Alpha"])
    ws.append(["B", "A", 1, 3, 5, "Beta"])
    xlsx_file = tmp_path / "pert_tasks.xlsx"
    wb.save(str(xlsx_file))

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-xlsx").set_input_files(str(xlsx_file))

    expect(page.locator("#toggle-pert")).to_be_checked()

    rows = page.locator("#input-table tbody tr")
    expect(rows).to_have_count(2)

    row1 = rows.nth(0)
    expect(row1.locator("td:nth-child(5)")).to_be_visible()
    expect(row1.locator("td:nth-child(5)")).to_have_text("2")
    expect(row1.locator("td:nth-child(6)")).to_have_text("4")
    expect(row1.locator("td:nth-child(7)")).to_have_text("6")

    row2 = rows.nth(1)
    expect(row2.locator("td:nth-child(5)")).to_have_text("1")
    expect(row2.locator("td:nth-child(6)")).to_have_text("3")
    expect(row2.locator("td:nth-child(7)")).to_have_text("5")


def test_xlsx_wrong_headers_shows_error_banner(page, tmp_path):
    """Uploading an Excel file with unrecognised column names shows an explicit column error."""
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.append(["id", "name", "duration", "description"])  # wrong — expects ac, pr, du
    ws.append(["A", "-", 3, "First task"])
    xlsx_file = tmp_path / "bad_headers.xlsx"
    wb.save(str(xlsx_file))

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-xlsx").set_input_files(str(xlsx_file))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("Failed to import Excel file")
    expect(page.locator("#out-text")).to_contain_text("Unrecognised column")


# ── Group 5: Error handling ───────────────────────────────────────────────────

def test_json_import_invalid_syntax_shows_error_banner(page, tmp_path):
    """Uploading a .json file with non-JSON content shows the error banner."""
    bad_file = tmp_path / "bad.json"
    bad_file.write_text("this is not json!!!", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(bad_file))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("Failed to import JSON")


def test_json_import_wrong_type_shows_error_banner(page, tmp_path):
    """Uploading a .json file that is not an array shows the error banner."""
    bad_file = tmp_path / "obj.json"
    bad_file.write_text('{"id": "A", "duration": 3}', encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(bad_file))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("array")


def test_json_import_unknown_fields_shows_error(page, tmp_path):
    """JSON with unrecognised field names shows an explicit field error."""
    bad_tasks = [{"identifier": "A", "time": 3, "deps": ""}]
    json_file = tmp_path / "bad_fields.json"
    json_file.write_text(json.dumps(bad_tasks), encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-json").set_input_files(str(json_file))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("Unrecognised field")


def test_csv_import_missing_header_shows_error_banner(page, tmp_path):
    """Uploading a CSV with only one line (no header) shows the error banner."""
    bad_csv = tmp_path / "bad.csv"
    bad_csv.write_text("A,-,3,Design\n", encoding="utf-8")

    page.goto(BASE_URL, wait_until="domcontentloaded")
    page.locator("#file-upload-csv").set_input_files(str(bad_csv))

    out = page.locator("#out")
    expect(out).to_be_visible()
    expect(out).to_have_class(re.compile(r"error"))
    expect(page.locator("#out-text")).to_contain_text("Failed to import CSV")


# ── Group 6: PNG exports ──────────────────────────────────────────────────────

def test_gantt_export_png_triggers_download_with_correct_filename(analyzed_page):
    """Clicking Export PNG on the Gantt tab triggers a download named gantt-chart.png."""
    page = analyzed_page
    page.locator("#gantt-tab").click()
    page.wait_for_selector(".gantt-row")

    with page.expect_download() as dl_info:
        page.locator("#btn-export-gantt").click()

    assert dl_info.value.suggested_filename == "gantt-chart.png"


def test_network_export_png_triggers_download_with_correct_filename(analyzed_page):
    """Clicking Export PNG on the Network tab triggers a download named network-diagram.png."""
    page = analyzed_page
    page.locator("#network-tab").click()
    expect(page.locator("#tab-network")).to_have_class(re.compile(r"active"))

    with page.expect_download() as dl_info:
        page.locator("#btn-export-network").click()

    assert dl_info.value.suggested_filename == "network-diagram.png"
